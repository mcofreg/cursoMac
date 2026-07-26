#!/usr/bin/env python3
"""Exporta el listado a Excel con filtros, anchos y paneles fijos.

El CSV es la fuente de verdad; esto es sólo una vista cómoda para trabajarlo.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pymes import pipeline

HOJAS = [
    ("Completo (5 campos)", "pymes_chilenas_completo.csv"),
    ("RUT confirmado", "pymes_chilenas_sin_sitio_web.csv"),
    ("Con Instagram", "pymes_con_redes_sin_sitio_web.csv"),
]
ANCHOS = {"nombre_empresa": 34, "razon_social": 34, "rut": 15, "confianza_rut": 20,
          "telefono": 16, "email": 32, "instagram": 38, "facebook": 38,
          "rubro": 20, "rubros": 26, "comuna": 18, "direccion": 34, "fuente": 30}

wb = Workbook()
wb.remove(wb.active)
cab_fill = PatternFill("solid", fgColor="1F3864")
cab_font = Font(color="FFFFFF", bold=True)

for titulo, archivo in HOJAS:
    ruta = pipeline.DATOS / archivo
    if not ruta.exists():
        continue
    filas = list(csv.DictReader(ruta.open(encoding="utf-8-sig")))
    if not filas:
        continue
    ws = wb.create_sheet(titulo[:31])
    cols = list(filas[0].keys())
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cel = ws.cell(row=1, column=c)
        cel.fill, cel.font = cab_fill, cab_font
        cel.alignment = Alignment(vertical="center")
    for f in filas:
        ws.append([f.get(c, "") for c in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(filas)+1}"
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS.get(c, 14)
    print(f"  {titulo}: {len(filas)} filas")

destino = pipeline.DATOS / "pymes_chilenas.xlsx"
wb.save(destino)
print(f"-> {destino}")
